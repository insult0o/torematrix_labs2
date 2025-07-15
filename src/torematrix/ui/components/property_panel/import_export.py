"""Property import/export functionality for data persistence and sharing"""

from typing import Dict, List, Any, Optional, Union, IO
from dataclasses import dataclass, asdict
from enum import Enum
from pathlib import Path
import json
import csv
import xml.etree.ElementTree as ET
from xml.dom import minidom
import yaml
import pickle
from datetime import datetime
from io import StringIO

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel, QPushButton,
    QComboBox, QCheckBox, QProgressBar, QTextEdit, QFileDialog, QMessageBox,
    QDialog, QDialogButtonBox, QTabWidget, QTableWidget, QTableWidgetItem,
    QHeaderView, QSplitter, QTreeWidget, QTreeWidgetItem
)
from PyQt6.QtCore import Qt, pyqtSignal, QThread, QObject, QTimer
from PyQt6.QtGui import QFont, QIcon

from .models import PropertyMetadata, PropertyValue, PropertyChange
from .events import PropertyNotificationCenter


class ExportFormat(Enum):
    """Supported export formats"""
    JSON = "json"
    CSV = "csv"
    XML = "xml"
    YAML = "yaml"
    PICKLE = "pickle"
    EXCEL = "xlsx"
    
    @property
    def extension(self) -> str:
        """Get file extension for format"""
        extensions = {
            ExportFormat.JSON: ".json",
            ExportFormat.CSV: ".csv",
            ExportFormat.XML: ".xml",
            ExportFormat.YAML: ".yaml",
            ExportFormat.PICKLE: ".pkl",
            ExportFormat.EXCEL: ".xlsx"
        }
        return extensions[self]
    
    @property
    def description(self) -> str:
        """Get human readable description"""
        descriptions = {
            ExportFormat.JSON: "JSON - JavaScript Object Notation",
            ExportFormat.CSV: "CSV - Comma Separated Values",
            ExportFormat.XML: "XML - Extensible Markup Language", 
            ExportFormat.YAML: "YAML - Yet Another Markup Language",
            ExportFormat.PICKLE: "Pickle - Python Binary Format",
            ExportFormat.EXCEL: "Excel - Microsoft Excel Format"
        }
        return descriptions[self]


@dataclass
class ExportConfiguration:
    """Configuration for export operations"""
    format: ExportFormat
    include_metadata: bool = True
    include_history: bool = False
    include_validation: bool = True
    flatten_nested: bool = False
    export_empty_values: bool = False
    custom_fields: Dict[str, Any] = None
    
    def __post_init__(self):
        if self.custom_fields is None:
            self.custom_fields = {}


@dataclass
class ImportConfiguration:
    """Configuration for import operations"""
    format: ExportFormat
    merge_strategy: str = "replace"  # replace, merge, skip
    validate_on_import: bool = True
    create_backup: bool = True
    ignore_errors: bool = False
    custom_mapping: Dict[str, str] = None
    
    def __post_init__(self):
        if self.custom_mapping is None:
            self.custom_mapping = {}


@dataclass
class ExportResult:
    """Result of export operation"""
    success: bool
    file_path: Optional[str] = None
    exported_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    errors: List[str] = None
    export_time: float = 0.0
    file_size: int = 0
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []


@dataclass
class ImportResult:
    """Result of import operation"""
    success: bool
    imported_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    errors: List[str] = None
    import_time: float = 0.0
    backup_path: Optional[str] = None
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []


class PropertyExporter(QObject):
    """Handles property data export to various formats"""
    
    # Signals
    export_started = pyqtSignal(str)  # format
    export_progress = pyqtSignal(int, str)  # percentage, status
    export_completed = pyqtSignal(object)  # ExportResult
    export_error = pyqtSignal(str)  # error message
    
    def __init__(self):
        super().__init__()
        self.property_manager = None
        self.cancel_requested = False
    
    def set_property_manager(self, manager) -> None:
        """Set property manager for data access"""
        self.property_manager = manager
    
    def export_properties(self, 
                         element_ids: List[str],
                         config: ExportConfiguration,
                         file_path: str) -> ExportResult:
        """Export properties to file"""
        self.cancel_requested = False
        start_time = datetime.now()
        
        try:
            self.export_started.emit(config.format.value)
            
            # Collect property data
            self.export_progress.emit(10, "Collecting property data...")
            property_data = self._collect_property_data(element_ids, config)
            
            # Export to format
            self.export_progress.emit(50, f"Exporting to {config.format.value.upper()}...")
            
            if config.format == ExportFormat.JSON:
                self._export_json(property_data, file_path, config)
            elif config.format == ExportFormat.CSV:
                self._export_csv(property_data, file_path, config)
            elif config.format == ExportFormat.XML:
                self._export_xml(property_data, file_path, config)
            elif config.format == ExportFormat.YAML:
                self._export_yaml(property_data, file_path, config)
            elif config.format == ExportFormat.PICKLE:
                self._export_pickle(property_data, file_path, config)
            elif config.format == ExportFormat.EXCEL:
                self._export_excel(property_data, file_path, config)
            
            self.export_progress.emit(90, "Finalizing export...")
            
            # Create result
            end_time = datetime.now()
            file_size = Path(file_path).stat().st_size if Path(file_path).exists() else 0
            
            result = ExportResult(
                success=True,
                file_path=file_path,
                exported_count=len(element_ids),
                export_time=(end_time - start_time).total_seconds(),
                file_size=file_size
            )
            
            self.export_progress.emit(100, "Export completed")
            self.export_completed.emit(result)
            
            return result
            
        except Exception as e:
            error_msg = f"Export failed: {str(e)}"
            self.export_error.emit(error_msg)
            
            return ExportResult(
                success=False,
                errors=[error_msg],
                export_time=(datetime.now() - start_time).total_seconds()
            )
    
    def _collect_property_data(self, element_ids: List[str], config: ExportConfiguration) -> Dict[str, Any]:
        """Collect property data from elements"""
        export_data = {
            'export_info': {
                'timestamp': datetime.now().isoformat(),
                'format': config.format.value,
                'element_count': len(element_ids),
                'configuration': asdict(config)
            },
            'elements': {}
        }
        
        for i, element_id in enumerate(element_ids):
            if self.cancel_requested:
                break
            
            # Update progress
            progress = int(10 + (i / len(element_ids)) * 40)
            self.export_progress.emit(progress, f"Processing element {i+1}/{len(element_ids)}")
            
            element_data = self._collect_element_data(element_id, config)
            export_data['elements'][element_id] = element_data
        
        return export_data
    
    def _collect_element_data(self, element_id: str, config: ExportConfiguration) -> Dict[str, Any]:
        """Collect data for single element"""
        if not self.property_manager:
            return {}
        
        element_data = {
            'properties': {},
            'metadata': {},
            'history': [],
            'validation': {}
        }
        
        # Get properties
        properties = self.property_manager.get_element_properties(element_id)
        for prop_name, prop_value in properties.items():
            if not config.export_empty_values and not prop_value:
                continue
            
            element_data['properties'][prop_name] = prop_value
            
            # Include metadata if requested
            if config.include_metadata:
                metadata = self.property_manager.get_property_metadata(element_id, prop_name)
                if metadata:
                    element_data['metadata'][prop_name] = asdict(metadata)
            
            # Include validation if requested
            if config.include_validation:
                validation_result = self.property_manager.validate_property(element_id, prop_name)
                if validation_result:
                    element_data['validation'][prop_name] = asdict(validation_result)
        
        # Include history if requested
        if config.include_history:
            history = self.property_manager.get_property_history(element_id)
            element_data['history'] = [asdict(change) for change in history]
        
        return element_data
    
    def _export_json(self, data: Dict[str, Any], file_path: str, config: ExportConfiguration) -> None:
        """Export to JSON format"""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    
    def _export_csv(self, data: Dict[str, Any], file_path: str, config: ExportConfiguration) -> None:
        """Export to CSV format"""
        elements = data.get('elements', {})
        
        # Flatten data for CSV
        rows = []
        headers = set(['element_id'])
        
        for element_id, element_data in elements.items():
            row = {'element_id': element_id}
            properties = element_data.get('properties', {})
            
            if config.flatten_nested:
                flattened = self._flatten_dict(properties)
                row.update(flattened)
                headers.update(flattened.keys())
            else:
                row.update(properties)
                headers.update(properties.keys())
            
            rows.append(row)
        
        # Write CSV
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            if rows:
                writer = csv.DictWriter(f, fieldnames=sorted(headers))
                writer.writeheader()
                writer.writerows(rows)
    
    def _export_xml(self, data: Dict[str, Any], file_path: str, config: ExportConfiguration) -> None:
        """Export to XML format"""
        root = ET.Element('property_export')
        
        # Add export info
        info_elem = ET.SubElement(root, 'export_info')
        for key, value in data.get('export_info', {}).items():
            child = ET.SubElement(info_elem, key)
            child.text = str(value)
        
        # Add elements
        elements_elem = ET.SubElement(root, 'elements')
        for element_id, element_data in data.get('elements', {}).items():
            element_elem = ET.SubElement(elements_elem, 'element')
            element_elem.set('id', element_id)
            
            # Add properties
            props_elem = ET.SubElement(element_elem, 'properties')
            for prop_name, prop_value in element_data.get('properties', {}).items():
                prop_elem = ET.SubElement(props_elem, 'property')
                prop_elem.set('name', prop_name)
                prop_elem.text = str(prop_value)
        
        # Pretty print XML
        xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(xml_str)
    
    def _export_yaml(self, data: Dict[str, Any], file_path: str, config: ExportConfiguration) -> None:
        """Export to YAML format"""
        with open(file_path, 'w', encoding='utf-8') as f:
            yaml.dump(data, f, default_flow_style=False, allow_unicode=True)
    
    def _export_pickle(self, data: Dict[str, Any], file_path: str, config: ExportConfiguration) -> None:
        """Export to Pickle format"""
        with open(file_path, 'wb') as f:
            pickle.dump(data, f)
    
    def _export_excel(self, data: Dict[str, Any], file_path: str, config: ExportConfiguration) -> None:
        """Export to Excel format"""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise ImportError("openpyxl library required for Excel export")
        
        wb = Workbook()
        
        # Properties sheet
        ws = wb.active
        ws.title = "Properties"
        
        elements = data.get('elements', {})
        if elements:
            # Create headers
            headers = ['Element ID']
            first_element = next(iter(elements.values()))
            properties = first_element.get('properties', {})
            headers.extend(sorted(properties.keys()))
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row, (element_id, element_data) in enumerate(elements.items(), 2):
                ws.cell(row=row, column=1, value=element_id)
                
                properties = element_data.get('properties', {})
                for col, prop_name in enumerate(sorted(properties.keys()), 2):
                    ws.cell(row=row, column=col, value=str(properties.get(prop_name, '')))
        
        # Metadata sheet if included
        if config.include_metadata:
            ws_meta = wb.create_sheet("Metadata")
            # TODO: Add metadata sheet implementation
        
        wb.save(file_path)
    
    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = '', sep: str = '.') -> Dict[str, Any]:
        """Flatten nested dictionary"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)
    
    def cancel_export(self) -> None:
        """Cancel current export operation"""
        self.cancel_requested = True


class PropertyImporter(QObject):
    """Handles property data import from various formats"""
    
    # Signals
    import_started = pyqtSignal(str)  # format
    import_progress = pyqtSignal(int, str)  # percentage, status
    import_completed = pyqtSignal(object)  # ImportResult
    import_error = pyqtSignal(str)  # error message
    
    def __init__(self):
        super().__init__()
        self.property_manager = None
        self.cancel_requested = False
    
    def set_property_manager(self, manager) -> None:
        """Set property manager for data writing"""
        self.property_manager = manager
    
    def import_properties(self, 
                         file_path: str,
                         config: ImportConfiguration) -> ImportResult:
        """Import properties from file"""
        self.cancel_requested = False
        start_time = datetime.now()
        
        try:
            self.import_started.emit(config.format.value)
            
            # Create backup if requested
            backup_path = None
            if config.create_backup:
                self.import_progress.emit(5, "Creating backup...")
                backup_path = self._create_backup()
            
            # Load data from file
            self.import_progress.emit(20, f"Loading {config.format.value.upper()} file...")
            data = self._load_data(file_path, config.format)
            
            # Process import
            self.import_progress.emit(40, "Processing import data...")
            result = self._process_import(data, config)
            
            # Finalize
            end_time = datetime.now()
            result.import_time = (end_time - start_time).total_seconds()
            result.backup_path = backup_path
            
            self.import_progress.emit(100, "Import completed")
            self.import_completed.emit(result)
            
            return result
            
        except Exception as e:
            error_msg = f"Import failed: {str(e)}"
            self.import_error.emit(error_msg)
            
            return ImportResult(
                success=False,
                errors=[error_msg],
                import_time=(datetime.now() - start_time).total_seconds()
            )
    
    def _load_data(self, file_path: str, format: ExportFormat) -> Dict[str, Any]:
        """Load data from file based on format"""
        if format == ExportFormat.JSON:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        
        elif format == ExportFormat.YAML:
            with open(file_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        
        elif format == ExportFormat.PICKLE:
            with open(file_path, 'rb') as f:
                return pickle.load(f)
        
        elif format == ExportFormat.CSV:
            return self._load_csv(file_path)
        
        elif format == ExportFormat.XML:
            return self._load_xml(file_path)
        
        elif format == ExportFormat.EXCEL:
            return self._load_excel(file_path)
        
        else:
            raise ValueError(f"Unsupported import format: {format}")
    
    def _load_csv(self, file_path: str) -> Dict[str, Any]:
        """Load data from CSV file"""
        elements = {}
        
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                element_id = row.pop('element_id', None)
                if element_id:
                    elements[element_id] = {
                        'properties': {k: v for k, v in row.items() if v}
                    }
        
        return {
            'export_info': {'format': 'csv'},
            'elements': elements
        }
    
    def _load_xml(self, file_path: str) -> Dict[str, Any]:
        """Load data from XML file"""
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        elements = {}
        elements_elem = root.find('elements')
        
        if elements_elem is not None:
            for element_elem in elements_elem.findall('element'):
                element_id = element_elem.get('id')
                if element_id:
                    properties = {}
                    props_elem = element_elem.find('properties')
                    
                    if props_elem is not None:
                        for prop_elem in props_elem.findall('property'):
                            prop_name = prop_elem.get('name')
                            prop_value = prop_elem.text
                            if prop_name:
                                properties[prop_name] = prop_value
                    
                    elements[element_id] = {'properties': properties}
        
        return {
            'export_info': {'format': 'xml'},
            'elements': elements
        }
    
    def _load_excel(self, file_path: str) -> Dict[str, Any]:
        """Load data from Excel file"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl library required for Excel import")
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        elements = {}
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Process data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            
            element_id = row[0] if row else None
            if element_id:
                properties = {}
                for i, value in enumerate(row[1:], 1):
                    if i < len(headers) and value is not None:
                        prop_name = headers[i]
                        properties[prop_name] = str(value)
                
                elements[str(element_id)] = {'properties': properties}
        
        return {
            'export_info': {'format': 'excel'},
            'elements': elements
        }
    
    def _process_import(self, data: Dict[str, Any], config: ImportConfiguration) -> ImportResult:
        """Process imported data"""
        result = ImportResult(success=True)
        elements = data.get('elements', {})
        total_elements = len(elements)
        
        for i, (element_id, element_data) in enumerate(elements.items()):
            if self.cancel_requested:
                break
            
            # Update progress
            progress = int(40 + (i / total_elements) * 50)
            self.import_progress.emit(progress, f"Importing element {i+1}/{total_elements}")
            
            try:
                success = self._import_element(element_id, element_data, config)
                if success:
                    result.imported_count += 1
                else:
                    result.skipped_count += 1
            
            except Exception as e:
                result.error_count += 1
                result.errors.append(f"Error importing {element_id}: {str(e)}")
                
                if not config.ignore_errors:
                    result.success = False
                    break
        
        return result
    
    def _import_element(self, element_id: str, element_data: Dict[str, Any], 
                       config: ImportConfiguration) -> bool:
        """Import single element"""
        if not self.property_manager:
            return False
        
        properties = element_data.get('properties', {})
        
        for prop_name, prop_value in properties.items():
            # Apply custom mapping if provided
            mapped_name = config.custom_mapping.get(prop_name, prop_name)
            
            # Handle merge strategy
            if config.merge_strategy == 'skip':
                current_value = self.property_manager.get_property_value(element_id, mapped_name)
                if current_value is not None:
                    continue
            
            # Validate if requested
            if config.validate_on_import:
                validation_result = self.property_manager.validate_property_value(
                    element_id, mapped_name, prop_value
                )
                if not validation_result.is_valid:
                    continue
            
            # Set property value
            self.property_manager.set_property_value(element_id, mapped_name, prop_value)
        
        return True
    
    def _create_backup(self) -> str:
        """Create backup of current data"""
        if not self.property_manager:
            return None
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"property_backup_{timestamp}.json"
        
        # TODO: Implement backup creation
        # This would export current data to backup file
        
        return backup_path
    
    def cancel_import(self) -> None:
        """Cancel current import operation"""
        self.cancel_requested = True


class ImportExportDialog(QDialog):
    """Dialog for import/export operations"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Import/Export Properties")
        self.setModal(True)
        self.resize(600, 500)
        
        self.exporter = PropertyExporter()
        self.importer = PropertyImporter()
        
        self._setup_ui()
        self._setup_connections()
    
    def _setup_ui(self) -> None:
        """Setup dialog UI"""
        layout = QVBoxLayout(self)
        
        # Tab widget for import/export
        self.tab_widget = QTabWidget()
        
        # Export tab
        export_tab = self._create_export_tab()
        self.tab_widget.addTab(export_tab, "Export")
        
        # Import tab
        import_tab = self._create_import_tab()
        self.tab_widget.addTab(import_tab, "Import")
        
        layout.addWidget(self.tab_widget)
        
        # Button box
        self.button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Close
        )
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)
    
    def _create_export_tab(self) -> QWidget:
        """Create export tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Format selection
        format_group = QGroupBox("Export Format")
        format_layout = QVBoxLayout(format_group)
        
        self.export_format_combo = QComboBox()
        for fmt in ExportFormat:
            self.export_format_combo.addItem(fmt.description, fmt)
        format_layout.addWidget(self.export_format_combo)
        
        layout.addWidget(format_group)
        
        # Options
        options_group = QGroupBox("Export Options")
        options_layout = QVBoxLayout(options_group)
        
        self.include_metadata_cb = QCheckBox("Include metadata")
        self.include_metadata_cb.setChecked(True)
        options_layout.addWidget(self.include_metadata_cb)
        
        self.include_history_cb = QCheckBox("Include change history")
        options_layout.addWidget(self.include_history_cb)
        
        self.flatten_nested_cb = QCheckBox("Flatten nested properties")
        options_layout.addWidget(self.flatten_nested_cb)
        
        layout.addWidget(options_group)
        
        # Progress
        self.export_progress = QProgressBar()
        self.export_progress.setVisible(False)
        layout.addWidget(self.export_progress)
        
        # Controls
        controls_layout = QHBoxLayout()
        
        self.export_btn = QPushButton("Export Properties")
        self.export_btn.clicked.connect(self._start_export)
        
        self.export_cancel_btn = QPushButton("Cancel")
        self.export_cancel_btn.clicked.connect(self._cancel_export)
        self.export_cancel_btn.setVisible(False)
        
        controls_layout.addWidget(self.export_btn)
        controls_layout.addWidget(self.export_cancel_btn)
        controls_layout.addStretch()
        
        layout.addLayout(controls_layout)
        
        return tab
    
    def _create_import_tab(self) -> QWidget:
        """Create import tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # File selection
        file_group = QGroupBox("Import File")
        file_layout = QHBoxLayout(file_group)
        
        self.import_file_label = QLabel("No file selected")
        file_layout.addWidget(self.import_file_label)
        
        self.browse_btn = QPushButton("Browse...")
        self.browse_btn.clicked.connect(self._browse_import_file)
        file_layout.addWidget(self.browse_btn)
        
        layout.addWidget(file_group)
        
        # Options
        options_group = QGroupBox("Import Options")
        options_layout = QVBoxLayout(options_group)
        
        self.merge_strategy_combo = QComboBox()
        self.merge_strategy_combo.addItems(["Replace existing", "Merge with existing", "Skip existing"])
        options_layout.addWidget(QLabel("Merge Strategy:"))
        options_layout.addWidget(self.merge_strategy_combo)
        
        self.validate_import_cb = QCheckBox("Validate on import")
        self.validate_import_cb.setChecked(True)
        options_layout.addWidget(self.validate_import_cb)
        
        self.create_backup_cb = QCheckBox("Create backup before import")
        self.create_backup_cb.setChecked(True)
        options_layout.addWidget(self.create_backup_cb)
        
        layout.addWidget(options_group)
        
        # Progress
        self.import_progress = QProgressBar()
        self.import_progress.setVisible(False)
        layout.addWidget(self.import_progress)
        
        # Controls
        controls_layout = QHBoxLayout()
        
        self.import_btn = QPushButton("Import Properties")
        self.import_btn.clicked.connect(self._start_import)
        self.import_btn.setEnabled(False)
        
        self.import_cancel_btn = QPushButton("Cancel")
        self.import_cancel_btn.clicked.connect(self._cancel_import)
        self.import_cancel_btn.setVisible(False)
        
        controls_layout.addWidget(self.import_btn)
        controls_layout.addWidget(self.import_cancel_btn)
        controls_layout.addStretch()
        
        layout.addLayout(controls_layout)
        
        return tab
    
    def _setup_connections(self) -> None:
        """Setup signal connections"""
        # Export connections
        self.exporter.export_progress.connect(self.export_progress.setValue)
        self.exporter.export_completed.connect(self._on_export_completed)
        self.exporter.export_error.connect(self._on_export_error)
        
        # Import connections
        self.importer.import_progress.connect(self.import_progress.setValue)
        self.importer.import_completed.connect(self._on_import_completed)
        self.importer.import_error.connect(self._on_import_error)
    
    def set_property_manager(self, manager) -> None:
        """Set property manager"""
        self.exporter.set_property_manager(manager)
        self.importer.set_property_manager(manager)
    
    def _start_export(self) -> None:
        """Start export operation"""
        # Get export configuration
        format = self.export_format_combo.currentData()
        config = ExportConfiguration(
            format=format,
            include_metadata=self.include_metadata_cb.isChecked(),
            include_history=self.include_history_cb.isChecked(),
            flatten_nested=self.flatten_nested_cb.isChecked()
        )
        
        # Get file path
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Properties",
            f"properties{format.extension}",
            f"{format.description} (*{format.extension})"
        )
        
        if file_path:
            # TODO: Get selected element IDs from property manager
            element_ids = []  # Would get from selection
            
            # Start export
            self.export_progress.setVisible(True)
            self.export_btn.setEnabled(False)
            self.export_cancel_btn.setVisible(True)
            
            self.exporter.export_properties(element_ids, config, file_path)
    
    def _cancel_export(self) -> None:
        """Cancel export operation"""
        self.exporter.cancel_export()
    
    def _browse_import_file(self) -> None:
        """Browse for import file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Properties",
            "",
            "All supported (*.json *.csv *.xml *.yaml *.pkl *.xlsx);;JSON files (*.json);;CSV files (*.csv);;XML files (*.xml);;YAML files (*.yaml);;Pickle files (*.pkl);;Excel files (*.xlsx)"
        )
        
        if file_path:
            self.import_file_label.setText(Path(file_path).name)
            self.import_file_label.setProperty('file_path', file_path)
            self.import_btn.setEnabled(True)
    
    def _start_import(self) -> None:
        """Start import operation"""
        file_path = self.import_file_label.property('file_path')
        if not file_path:
            return
        
        # Detect format from file extension
        suffix = Path(file_path).suffix.lower()
        format_map = {
            '.json': ExportFormat.JSON,
            '.csv': ExportFormat.CSV,
            '.xml': ExportFormat.XML,
            '.yaml': ExportFormat.YAML,
            '.yml': ExportFormat.YAML,
            '.pkl': ExportFormat.PICKLE,
            '.xlsx': ExportFormat.EXCEL
        }
        
        format = format_map.get(suffix)
        if not format:
            QMessageBox.warning(self, "Unsupported Format", f"Unsupported file format: {suffix}")
            return
        
        # Get import configuration
        merge_strategies = ['replace', 'merge', 'skip']
        merge_strategy = merge_strategies[self.merge_strategy_combo.currentIndex()]
        
        config = ImportConfiguration(
            format=format,
            merge_strategy=merge_strategy,
            validate_on_import=self.validate_import_cb.isChecked(),
            create_backup=self.create_backup_cb.isChecked()
        )
        
        # Start import
        self.import_progress.setVisible(True)
        self.import_btn.setEnabled(False)
        self.import_cancel_btn.setVisible(True)
        
        self.importer.import_properties(file_path, config)
    
    def _cancel_import(self) -> None:
        """Cancel import operation"""
        self.importer.cancel_import()
    
    def _on_export_completed(self, result: ExportResult) -> None:
        """Handle export completion"""
        self.export_progress.setVisible(False)
        self.export_btn.setEnabled(True)
        self.export_cancel_btn.setVisible(False)
        
        if result.success:
            QMessageBox.information(
                self,
                "Export Complete",
                f"Successfully exported {result.exported_count} elements to {result.file_path}"
            )
        else:
            QMessageBox.warning(
                self,
                "Export Failed",
                f"Export failed: {'; '.join(result.errors)}"
            )
    
    def _on_export_error(self, error_msg: str) -> None:
        """Handle export error"""
        self.export_progress.setVisible(False)
        self.export_btn.setEnabled(True)
        self.export_cancel_btn.setVisible(False)
        
        QMessageBox.critical(self, "Export Error", error_msg)
    
    def _on_import_completed(self, result: ImportResult) -> None:
        """Handle import completion"""
        self.import_progress.setVisible(False)
        self.import_btn.setEnabled(True)
        self.import_cancel_btn.setVisible(False)
        
        if result.success:
            message = f"Successfully imported {result.imported_count} elements"
            if result.updated_count > 0:
                message += f", updated {result.updated_count} elements"
            if result.skipped_count > 0:
                message += f", skipped {result.skipped_count} elements"
            
            QMessageBox.information(self, "Import Complete", message)
        else:
            QMessageBox.warning(
                self,
                "Import Failed",
                f"Import failed: {'; '.join(result.errors)}"
            )
    
    def _on_import_error(self, error_msg: str) -> None:
        """Handle import error"""
        self.import_progress.setVisible(False)
        self.import_btn.setEnabled(True)
        self.import_cancel_btn.setVisible(False)
        
        QMessageBox.critical(self, "Import Error", error_msg)


# Export import/export components
__all__ = [
    'ExportFormat',
    'ExportConfiguration',
    'ImportConfiguration',
    'ExportResult',
    'ImportResult',
    'PropertyExporter',
    'PropertyImporter',
    'ImportExportDialog'
]
